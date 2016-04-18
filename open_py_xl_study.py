# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.cell import get_column_letter
from openpyxl.styles import colors, Alignment, fills, PatternFill

FILE_NAME = 'output.xlsx'
SHEET_NAME = 'Sheet1'
START_COL = 1
TITLE_ROW = 1
DATA_ROW = 2
TITLE_LIST = ['ID', '名前', '年齢', '誕生日', '給料']
TALENT_LIST = [
    {
        "name": "田代まさお",
        "age": "59",
        "birthday": "1956/08/30",
        "salary": "1000000",
    },
    {
        "name": "志村けんご",
        "age": "65",
        "birthday": "1950/02/21",
        "salary": "50000000",
    },
    {
        "name": "鈴木雅",
        "age": "59",
        "birthday": "1956/09/23",
        "salary": "30000000",
    },
]

# シートの設定
wb = openpyxl.Workbook()

# アクティブなシートを取得
ws = wb.active

# シートに名前を設定
ws.title = SHEET_NAME


# スタイル設定
def write_title_cell(worksheet, row, column, value, font_color, fill_color):
    cell = worksheet.cell(row=row, column=column)
    cell.font = cell.font.copy(color=font_color)
    cell.fill = PatternFill(patternType=fills.FILL_SOLID, fgColor=fill_color)
    cell.alignment = Alignment(horizontal='center')
    cell.value = value


def write_cell(worksheet, row, column, value, number_format='General'):
    cell = worksheet.cell(row=row, column=column)
    cell.number_format = number_format
    cell.value = value


# タイトル行入力
col = START_COL
for title in TITLE_LIST:
    write_title_cell(worksheet=ws, row=TITLE_ROW, column=col, value=title, font_color=colors.WHITE,
                     fill_color=colors.DARKBLUE)

    # 表示幅を設定
    ws.column_dimensions[get_column_letter(col)].width = 13

    # カラムをインクリメント
    col += 1

# データ入力
data_row = DATA_ROW
for talent in TALENT_LIST:
    write_cell(worksheet=ws, row=data_row, column=1, value='=ROW()-1')
    write_cell(worksheet=ws, row=data_row, column=2, value=talent['name'])
    write_cell(worksheet=ws, row=data_row, column=3, value=int(talent['age']), number_format='#,##0')
    write_cell(worksheet=ws, row=data_row, column=4, value=talent['birthday'], number_format='yyyy年mm月dd日')
    write_cell(worksheet=ws, row=data_row, column=5, value=int(talent['salary']), number_format='"¥"#,##0')

    data_row += 1

# セルに値を入れる(合計行)
sum_target_row = ws.max_row
sum_row = len(TALENT_LIST) + 2

# タイトル設定
write_title_cell(worksheet=ws, row=sum_row, column=1, value='合計', font_color=colors.WHITE, fill_color=colors.RED)

# 合計行の設定
write_cell(worksheet=ws, row=sum_row, column=3, value='=SUM(C%d:C%d)' % (DATA_ROW, sum_target_row),
           number_format='#,##0')
write_cell(worksheet=ws, row=sum_row, column=5, value='=SUM(E%d:E%d)' % (DATA_ROW, sum_target_row),
           number_format='"¥"#,##0')

# ファイル出力
wb.save(filename=FILE_NAME)
